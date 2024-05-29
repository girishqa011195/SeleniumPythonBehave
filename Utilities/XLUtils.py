import openpyxl


def get_excel_value(parameter_id):
    file_location = "path/to/your/excel/file.xlsx"
    current_sheet_number = 0  # Index of the sheet (0-based)
    current_row = 1  # Specify the row from which to read data

    try:
        workbook = openpyxl.load_workbook(file_location)
        worksheet = workbook.worksheets[current_sheet_number]

        # Loop through each cell in the first row to find the column index matching the parameter ID
        j = 0
        while worksheet.cell(row=1, column=j + 1).value is not None:
            if str(worksheet.cell(row=1, column=j + 1).value).lower() == str(parameter_id).lower():
                # Get the value from the corresponding cell in the current row
                cell_value = worksheet.cell(row=current_row, column=j + 1).value
                if isinstance(cell_value, (int, float)):
                    return str(cell_value)  # Return numeric value as string
                else:
                    return cell_value  # Return string value
            j += 1

        # If parameter ID not found, return empty string
        return ""

    except FileNotFoundError:
        print(f"Excel file not found at location: {file_location}")
        return ""
    except Exception as e:
        print(f"An error occurred: {e}")
        return ""

def getRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_row)

def getColumnCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_column)

def readData(file,sheetName,rownum,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rownum, column=columnno).value

def writeData(file,sheetName,rownum,columnno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row=rownum, column=columnno).value = data
    workbook.save(file)
